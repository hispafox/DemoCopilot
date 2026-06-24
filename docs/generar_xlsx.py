"""Genera insights-proyecto.xlsx con estimaciones de trabajo del proyecto DemoCopilot."""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(__file__), "insights-proyecto.xlsx")

# ── Paleta ────────────────────────────────────────────────────────────────────
C_AZUL_OSC  = "1E2761"
C_AZUL_MED  = "2E75B6"
C_AZUL_CLAR = "CADCFC"
C_AZUL_PAL  = "EBF3FB"
C_VERDE     = "00B050"
C_NARANJA   = "FF8C00"
C_ROJO      = "C00000"
C_GRIS_CAB  = "404040"
C_GRIS_CLAR = "F4F4F4"
C_AMARILLO  = "FFFF00"
C_BLANCO    = "FFFFFF"

FONT_NOMBRE = "Arial"

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name=FONT_NOMBRE, bold=bold, size=size, color=color, italic=italic)

def alinear(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def borde_fino():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def borde_medio():
    s = Side(style="medium", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def aplicar(ws, cell_ref, value=None, bold=False, size=11, txt_color="000000",
            bg=None, h_align="left", wrap=False, italic=False, num_format=None, border=None):
    c = ws[cell_ref]
    if value is not None:
        c.value = value
    c.font      = font(bold=bold, size=size, color=txt_color, italic=italic)
    c.alignment = alinear(h_align, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if num_format:
        c.number_format = num_format
    if border:
        c.border = border
    return c


# ── Helpers de tabla ──────────────────────────────────────────────────────────
def cabecera_tabla(ws, row, cols_headers, col_start=1, bg=C_AZUL_MED):
    """Escribe una fila de cabecera y devuelve la fila siguiente."""
    for i, h in enumerate(cols_headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font      = font(bold=True, color=C_BLANCO)
        c.fill      = fill(bg)
        c.alignment = alinear("center")
        c.border    = borde_fino()
    return row + 1

def fila_datos(ws, row, values, col_start=1, alternada=False, bold=False,
               formats=None, aligns=None, txt_colors=None):
    bg = C_AZUL_PAL if alternada else C_BLANCO
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=v)
        c.font      = font(bold=bold, size=11,
                           color=txt_colors[i] if txt_colors else "000000")
        c.fill      = fill(bg)
        c.alignment = alinear(
            aligns[i] if aligns else ("right" if isinstance(v, (int, float)) else "left"),
            wrap=True)
        c.border    = borde_fino()
        if formats and formats[i]:
            c.number_format = formats[i]
    return row + 1

def titulo_seccion(ws, row, texto, col_start=1, col_end=10):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=texto)
    c.font      = font(bold=True, size=13, color=C_BLANCO)
    c.fill      = fill(C_AZUL_OSC)
    c.alignment = alinear("left")
    return row + 1

def fila_total(ws, row, label, formula, col_start=1, num_cols=8,
               label_col=None, formula_col=None, bg=C_AZUL_CLAR):
    lc = label_col   or col_start
    fc = formula_col or col_start + num_cols - 1
    for col in range(col_start, col_start + num_cols):
        c = ws.cell(row=row, column=col)
        c.fill   = fill(bg)
        c.border = borde_fino()
    ws.cell(row=row, column=lc, value=label).font  = font(bold=True)
    ws.cell(row=row, column=fc, value=formula).font = font(bold=True, color="000000")
    ws.cell(row=row, column=fc).number_format = '#,##0.0'
    return row + 1


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — Resumen ejecutivo
# ══════════════════════════════════════════════════════════════════════════════
def hoja_resumen(wb):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    # Cabecera principal
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "DemoCopilot — Estimación de Esfuerzo del Proyecto"
    c.font      = font(bold=True, size=18, color=C_BLANCO)
    c.fill      = fill(C_AZUL_OSC)
    c.alignment = alinear("center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = "Lista de Tareas · ASP.NET Core 8 · GitHub Copilot Demo"
    c.font      = font(italic=True, size=11, color=C_BLANCO)
    c.fill      = fill(C_AZUL_MED)
    c.alignment = alinear("center")
    ws.row_dimensions[2].height = 20

    # Tarjetas de métricas (fila 4-7)
    metricas = [
        ("Total horas (sin Copilot)", "='Estimaciones'!J3",  C_ROJO,    "A"),
        ("Total horas (con Copilot)", "='Estimaciones'!K3",  C_VERDE,   "C"),
        ("Ahorro estimado (horas)",   "=A5-C5",              C_NARANJA, "E"),
        ("Ahorro (%)",                "=(A5-C5)/A5",         C_AZUL_MED,"G"),
    ]

    for label, formula, color, col in metricas:
        ws.merge_cells(f"{col}4:{chr(ord(col)+1)}4")
        ws.merge_cells(f"{col}5:{chr(ord(col)+1)}5")
        ws.merge_cells(f"{col}6:{chr(ord(col)+1)}6")
        lc = ws[f"{col}4"]
        lc.value     = label
        lc.font      = font(bold=True, size=10, color="404040")
        lc.fill      = fill(C_GRIS_CLAR)
        lc.alignment = alinear("center")
        lc.border    = borde_medio()
        vc = ws[f"{col}5"]
        vc.value     = formula
        vc.font      = font(bold=True, size=22, color=color)
        vc.fill      = fill(C_BLANCO)
        vc.alignment = alinear("center")
        vc.border    = borde_medio()
        if "%" in label:
            vc.number_format = "0.0%"
        else:
            vc.number_format = '#,##0.0 "h"'
        ws[f"{col}6"].fill   = fill(C_BLANCO)
        ws[f"{col}6"].border = borde_medio()

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 40
    ws.row_dimensions[6].height = 8

    # Nota metodológica
    ws.merge_cells("A8:H8")
    ws["A8"].value     = "Metodología"
    ws["A8"].font      = font(bold=True, size=12, color=C_BLANCO)
    ws["A8"].fill      = fill(C_GRIS_CAB)
    ws["A8"].alignment = alinear("left")

    notas = [
        ("Estimación sin Copilot",  "Tiempo medio de un desarrollador .NET senior trabajando solo, con búsquedas en documentación, StackOverflow y prueba/error habituales."),
        ("Estimación con Copilot",  "Mismo desarrollador usando GitHub Copilot como asistente: autocompletado, generación de boilerplate, tests y explicaciones en contexto. Factor de aceleración ~2-3x en tareas repetitivas."),
        ("Unidad de medida",        "Horas-persona (h). Incluye implementación, tests unitarios y revisión. NO incluye reuniones ni gestión de proyecto."),
        ("Nivel de confianza",      "±20% en todas las estimaciones. Las tareas de infraestructura tienen mayor variabilidad que las de lógica de negocio."),
    ]
    row = 9
    for titulo, desc in notas:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"].value     = titulo
        ws[f"A{row}"].font      = font(bold=True, size=10)
        ws[f"A{row}"].fill      = fill(C_AZUL_PAL)
        ws[f"A{row}"].border    = borde_fino()
        ws.merge_cells(f"C{row}:H{row}")
        ws[f"C{row}"].value     = desc
        ws[f"C{row}"].font      = font(size=10)
        ws[f"C{row}"].alignment = alinear(wrap=True)
        ws[f"C{row}"].fill      = fill(C_BLANCO)
        ws[f"C{row}"].border    = borde_fino()
        ws.row_dimensions[row].height = 32
        row += 1

    # Anchos
    anchos = {"A": 22, "B": 22, "C": 22, "D": 6, "E": 22, "F": 6, "G": 22, "H": 6}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — Estimaciones detalladas
# ══════════════════════════════════════════════════════════════════════════════
def hoja_estimaciones(wb):
    ws = wb.create_sheet("Estimaciones")
    ws.sheet_view.showGridLines = False

    # Cabecera
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value     = "Estimaciones de Esfuerzo por Funcionalidad"
    c.font      = font(bold=True, size=15, color=C_BLANCO)
    c.fill      = fill(C_AZUL_MED)
    c.alignment = alinear("center")
    ws.row_dimensions[1].height = 28

    # Columnas
    headers = [
        "Área", "Funcionalidad", "Descripción",
        "Complejidad", "Prioridad",
        "h sin Copilot\n(min)", "h sin Copilot\n(prob)", "h sin Copilot\n(máx)",
        "h con Copilot\n(prob)",
        "Total sin Copilot\n(estimado)", "Total con Copilot\n(estimado)"
    ]

    # Fila 2: sub-cabecera grupos
    ws.merge_cells("F2:H2")
    ws["F2"].value     = "Sin GitHub Copilot"
    ws["F2"].font      = font(bold=True, color=C_BLANCO, size=10)
    ws["F2"].fill      = fill(C_ROJO)
    ws["F2"].alignment = alinear("center")
    ws.merge_cells("I2:K2")  # corregido: I-K para con Copilot
    ws["I2"].value     = "Con GitHub Copilot"
    ws["I2"].font      = font(bold=True, color=C_BLANCO, size=10)
    ws["I2"].fill      = fill(C_VERDE)
    ws["I2"].alignment = alinear("center")

    row = cabecera_tabla(ws, 3, headers, col_start=1, bg=C_AZUL_OSC)

    # Datos de funcionalidades
    # Formato columnas: Área, Funcionalidad, Descripción, Complejidad, Prioridad,
    #                   h_min, h_prob, h_max, h_copilot_prob, formula_sin, formula_con
    # Las fórmulas se añaden abajo
    funcionalidades = [
        # Infraestructura y setup
        ("Infraestructura", "Setup del proyecto",          "Crear solución .NET 8, configurar paquetes NuGet, estructura de carpetas",         "Baja",   "Alta",   1.0, 2.0, 3.0,  0.5),
        ("Infraestructura", "Configuración EF Core",       "AppDbContext, SQLite, cadena de conexión, Program.cs",                             "Media",  "Alta",   2.0, 3.0, 5.0,  1.0),
        ("Infraestructura", "Migraciones iniciales",       "Primera migración CreacionInicial + update database",                              "Baja",   "Alta",   0.5, 1.0, 1.5,  0.3),
        ("Infraestructura", "Datos de ejemplo (seeder)",   "DatosEjemplo.cs con registros representativos de cada entidad",                    "Baja",   "Media",  1.0, 1.5, 2.5,  0.5),
        ("Infraestructura", "Configuración de Swagger",    "Swagger/OpenAPI en Program.cs, anotaciones XML en controladores",                  "Baja",   "Media",  0.5, 1.0, 1.5,  0.3),
        # Modelos
        ("Modelos",         "TodoItem",                    "Entidad principal con todos los campos: repetitiva, recurrencia, FK opcionales",    "Media",  "Alta",   1.0, 1.5, 2.0,  0.4),
        ("Modelos",         "UsuarioAsignado",             "Entidad usuario con nombre y email, relación 1:N con TodoItem",                    "Baja",   "Alta",   0.5, 1.0, 1.5,  0.3),
        ("Modelos",         "PlantillaTarea",              "Entidad plantilla reutilizable con recurrencia por defecto",                        "Baja",   "Alta",   0.5, 1.0, 1.5,  0.3),
        ("Modelos",         "TipoRecurrencia (enum)",      "Enum Diaria/Semanal/Mensual",                                                      "Baja",   "Alta",   0.2, 0.3, 0.5,  0.1),
        ("Modelos",         "Fluent API / configuración",  "Configuración de relaciones, índices y restricciones en AppDbContext",              "Media",  "Alta",   1.5, 2.0, 3.0,  0.7),
        # DTOs
        ("DTOs",            "TareasDtos",                  "CrearTareaDto, ActualizarTareaDto, TareaResponseDto con validaciones básicas",      "Baja",   "Alta",   1.0, 1.5, 2.0,  0.4),
        ("DTOs",            "UsuariosAsignadosDtos",       "CrearUsuarioDto, ActualizarUsuarioDto, UsuarioResponseDto",                         "Baja",   "Alta",   0.5, 1.0, 1.5,  0.3),
        ("DTOs",            "PlantillasDtos",              "CrearPlantillaDto, ActualizarPlantillaDto, PlantillaResponseDto",                   "Baja",   "Alta",   0.5, 1.0, 1.5,  0.3),
        # Lógica de negocio
        ("Lógica Negocio",  "ITodoLogica + TodoLogica",    "CRUD básico de tareas + lógica de completar con generación de recurrencia",        "Alta",   "Alta",   3.0, 5.0, 8.0,  1.5),
        ("Lógica Negocio",  "IUsuarioAsignadoLogica",      "CRUD de usuarios, validación de email único",                                      "Media",  "Alta",   2.0, 3.0, 4.0,  1.0),
        ("Lógica Negocio",  "IPlantillaLogica",            "CRUD de plantillas + instanciar plantilla en nueva tarea",                         "Media",  "Alta",   2.0, 3.0, 5.0,  1.0),
        ("Lógica Negocio",  "Validaciones de negocio",     "Validar existencia por ID, reglas de recurrencia, email formato",                  "Media",  "Alta",   2.0, 3.0, 4.0,  1.0),
        # Servicios
        ("Servicios",       "ITodoService + TodoService",  "Mapeo DTO↔Entidad, orquestación de llamadas a lógica de negocio",                 "Media",  "Alta",   2.0, 3.0, 4.0,  1.0),
        ("Servicios",       "IUsuarioAsignadoService",     "Capa de servicio para usuarios, mapeos",                                           "Baja",   "Alta",   1.0, 1.5, 2.5,  0.5),
        ("Servicios",       "IPlantillaService",           "Capa de servicio para plantillas, mapeos",                                         "Baja",   "Alta",   1.0, 1.5, 2.5,  0.5),
        ("Servicios",       "Registro de DI en Program.cs","AddScoped para todos los servicios y lógicas, inyección de dependencias",         "Baja",   "Alta",   0.5, 0.5, 1.0,  0.2),
        # Controladores
        ("Controladores",   "TareasController",            "6 endpoints: GET, GET/{id}, POST, PUT, DELETE, POST/completar",                    "Media",  "Alta",   2.0, 3.0, 4.0,  0.8),
        ("Controladores",   "UsuariosAsignadosController", "5 endpoints CRUD estándar",                                                        "Baja",   "Alta",   1.5, 2.0, 3.0,  0.6),
        ("Controladores",   "PlantillasController",        "6 endpoints CRUD + instanciar",                                                    "Media",  "Alta",   2.0, 2.5, 3.5,  0.7),
        # Tests
        ("Tests",           "Tests TareasService",         "Tests unitarios CRUD + completar con y sin recurrencia (xUnit + Moq)",             "Alta",   "Alta",   3.0, 5.0, 8.0,  1.5),
        ("Tests",           "Tests UsuariosService",       "Tests unitarios CRUD + validación email único",                                    "Media",  "Media",  2.0, 3.0, 5.0,  1.0),
        ("Tests",           "Tests PlantillaService",      "Tests unitarios CRUD + instanciar",                                               "Media",  "Media",  2.0, 3.0, 5.0,  1.0),
        ("Tests",           "Tests controladores",         "Tests de integración básicos de los endpoints principales",                        "Alta",   "Baja",   4.0, 6.0, 10.0, 2.0),
        # Documentación
        ("Documentación",   "Análisis y Diseño (Markdown)","docs/analisis-diseño.md con arquitectura, modelos, endpoints",                    "Baja",   "Media",  1.0, 2.0, 3.0,  0.5),
        ("Documentación",   "README.md",                   "Instrucciones de arranque, estructura del proyecto, comandos útiles",              "Baja",   "Baja",   0.5, 1.0, 1.5,  0.3),
    ]

    # Colores de complejidad
    comp_colores = {"Baja": C_VERDE, "Media": C_NARANJA, "Alta": C_ROJO}
    prio_colores = {"Alta": C_ROJO, "Media": C_NARANJA, "Baja": "808080"}

    data_start_row = row
    alt = False
    for i, (area, func, desc, comp, prio, hmin, hprob, hmax, hcop) in enumerate(funcionalidades):
        excel_row = row
        bg = C_AZUL_PAL if alt else C_BLANCO
        alt = not alt

        vals = [area, func, desc, comp, prio, hmin, hprob, hmax, hcop]
        for j, v in enumerate(vals):
            col = j + 1
            c = ws.cell(row=excel_row, column=col, value=v)
            c.fill      = fill(bg)
            c.border    = borde_fino()
            c.alignment = alinear("left" if j < 5 else "center", wrap=(j == 2))
            if j < 5:
                c.font = font(size=10)
            else:
                c.font = font(size=10, color="0000FF")   # azul = inputs editables
                c.number_format = '#,##0.0'

            # Color complejidad/prioridad
            if j == 3:
                c.font = font(bold=True, color=comp_colores.get(v, "000000"), size=10)
            if j == 4:
                c.font = font(bold=True, color=prio_colores.get(v, "000000"), size=10)

        # Fórmulas columnas J (sin Copilot) y K (con Copilot)
        f_col  = get_column_letter(10)  # J
        ck_col = get_column_letter(11)  # K
        fj = ws.cell(row=excel_row, column=10,
                     value=f"=G{excel_row}")   # usa valor probable como estimado
        fj.fill         = fill(bg)
        fj.border       = borde_fino()
        fj.alignment    = alinear("center")
        fj.font         = font(size=10, color="000000")
        fj.number_format = '#,##0.0'

        fk = ws.cell(row=excel_row, column=11,
                     value=f"=I{excel_row}")
        fk.fill         = fill(bg)
        fk.border       = borde_fino()
        fk.alignment    = alinear("center")
        fk.font         = font(size=10, color="000000")
        fk.number_format = '#,##0.0'

        ws.row_dimensions[excel_row].height = 36
        row += 1

    data_end_row = row - 1

    # Fila TOTAL (fila 3 es donde están los sumatorios que referencia Resumen)
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL PROYECTO").font = font(bold=True, size=11, color=C_BLANCO)
    ws.cell(row=total_row, column=1).fill = fill(C_AZUL_OSC)
    ws.merge_cells(f"A{total_row}:E{total_row}")
    for col in range(1, 12):
        c = ws.cell(row=total_row, column=col)
        c.border = borde_fino()
        if col > 5:
            c.fill = fill(C_AZUL_OSC)
            c.font = font(bold=True, color=C_BLANCO, size=11)
            c.alignment = alinear("center")
            c.number_format = '#,##0.0 "h"'

    for col, formula_template in [
        (6,  f"=SUM(F{data_start_row}:F{data_end_row})"),
        (7,  f"=SUM(G{data_start_row}:G{data_end_row})"),
        (8,  f"=SUM(H{data_start_row}:H{data_end_row})"),
        (9,  f"=SUM(I{data_start_row}:I{data_end_row})"),
        (10, f"=SUM(J{data_start_row}:J{data_end_row})"),
        (11, f"=SUM(K{data_start_row}:K{data_end_row})"),
    ]:
        c = ws.cell(row=total_row, column=col, value=formula_template)
        c.fill         = fill(C_AZUL_OSC)
        c.font         = font(bold=True, color=C_BLANCO, size=11)
        c.alignment    = alinear("center")
        c.border       = borde_fino()
        c.number_format = '#,##0.0 "h"'

    # Fila 3 = fila de totales accesible desde Resumen (usamos named range via celda fija)
    # Ponemos los totales también en fila 3 para que el Resumen los lea directamente
    # Actualizo referencias en Resumen: usaremos la fila total_row
    # El resumen referencia ='Estimaciones'!J3 y K3 → los pongo en fila 3 como alias
    ws.cell(row=3, column=10, value=f"=J{total_row}").number_format = '#,##0.0'
    ws.cell(row=3, column=11, value=f"=K{total_row}").number_format = '#,##0.0'
    # Ocultar esas celdas de la fila 3 bajo la cabecera (ya están en fila de cabecera)

    ws.row_dimensions[total_row].height = 28

    # Nota sobre colores
    nota_row = total_row + 2
    ws.merge_cells(f"A{nota_row}:K{nota_row}")
    ws[f"A{nota_row}"].value = (
        "Leyenda:  🔵 Azul = inputs editables  |  "
        "Complejidad: Verde=Baja  Naranja=Media  Rojo=Alta  |  "
        "Estimaciones en horas-persona (h)  |  Confianza: ±20%"
    )
    ws[f"A{nota_row}"].font      = font(italic=True, size=9, color="606060")
    ws[f"A{nota_row}"].alignment = alinear("left")

    # Anchos de columna
    anchos = {
        "A": 18, "B": 28, "C": 42, "D": 13, "E": 11,
        "F": 14, "G": 14, "H": 14, "I": 16, "J": 18, "K": 18,
    }
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # Inmovilizar cabeceras
    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — Resumen por área
# ══════════════════════════════════════════════════════════════════════════════
def hoja_por_area(wb):
    ws = wb.create_sheet("Por Área")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "Resumen de Esfuerzo por Área"
    c.font      = font(bold=True, size=14, color=C_BLANCO)
    c.fill      = fill(C_AZUL_MED)
    c.alignment = alinear("center")
    ws.row_dimensions[1].height = 28

    row = cabecera_tabla(ws, 2, [
        "Área", "Funcionalidades",
        "Horas sin Copilot (prob)", "Horas con Copilot (prob)",
        "Ahorro (h)", "Ahorro (%)"
    ])

    areas = [
        ("Infraestructura",  5,  9.0,  2.8),
        ("Modelos",          5,  5.8,  1.8),
        ("DTOs",             3,  3.5,  1.0),
        ("Lógica Negocio",   4, 14.0,  4.5),
        ("Servicios",        4,  6.5,  2.2),
        ("Controladores",    3,  7.5,  2.1),
        ("Tests",            4, 17.0,  5.5),
        ("Documentación",    2,  3.0,  0.8),
    ]

    alt = False
    sum_sin_row_start = row
    for area, n_func, h_sin, h_cop in areas:
        ahorro_h   = f"=C{row}-D{row}"
        ahorro_pct = f"=IF(C{row}>0,(C{row}-D{row})/C{row},0)"
        bg = C_AZUL_PAL if alt else C_BLANCO
        alt = not alt

        data = [area, n_func, h_sin, h_cop, ahorro_h, ahorro_pct]
        fmts = [None, "#,##0", '#,##0.0 "h"', '#,##0.0 "h"', '#,##0.0 "h"', "0.0%"]
        for i, (v, fmt) in enumerate(zip(data, fmts)):
            c = ws.cell(row=row, column=i+1, value=v)
            c.fill         = fill(bg)
            c.border       = borde_fino()
            c.alignment    = alinear("center" if i > 0 else "left")
            c.font         = font(size=11,
                                  color="0000FF" if i in (2, 3) else "000000")
            if fmt:
                c.number_format = fmt
        row += 1

    sum_sin_row_end = row - 1

    # Total
    total_row = row
    ws.merge_cells(f"A{total_row}:B{total_row}")
    ws.cell(row=total_row, column=1, value="TOTAL").font  = font(bold=True, color=C_BLANCO)
    ws.cell(row=total_row, column=1).fill  = fill(C_AZUL_OSC)
    ws.cell(row=total_row, column=2).fill  = fill(C_AZUL_OSC)

    totales = [
        (3, f"=SUM(C{sum_sin_row_start}:C{sum_sin_row_end})", '#,##0.0 "h"'),
        (4, f"=SUM(D{sum_sin_row_start}:D{sum_sin_row_end})", '#,##0.0 "h"'),
        (5, f"=C{total_row}-D{total_row}",                    '#,##0.0 "h"'),
        (6, f"=IF(C{total_row}>0,(C{total_row}-D{total_row})/C{total_row},0)", "0.0%"),
    ]
    for col, formula, fmt in totales:
        c = ws.cell(row=total_row, column=col, value=formula)
        c.font         = font(bold=True, color=C_BLANCO)
        c.fill         = fill(C_AZUL_OSC)
        c.border       = borde_fino()
        c.alignment    = alinear("center")
        c.number_format = fmt

    ws.row_dimensions[total_row].height = 24

    # Equivalencia en jornadas
    nota_row = total_row + 2
    ws.merge_cells(f"A{nota_row}:F{nota_row}")
    ws.cell(row=nota_row, column=1).value = (
        f"Equivalencia:  1 jornada laboral = 8 h  ·  "
        f"Horas sin Copilot ≈  =ROUND(C{total_row}/8,1)  jornadas  ·  "
        f"Con Copilot ≈  =ROUND(D{total_row}/8,1)  jornadas"
    )
    ws.cell(row=nota_row, column=1).font      = font(italic=True, size=10)
    ws.cell(row=nota_row, column=1).alignment = alinear("left")

    # Fila de jornadas reales
    jornadas_row = nota_row + 1
    ws.cell(row=jornadas_row, column=1, value="Jornadas sin Copilot").font = font(bold=True)
    ws.cell(row=jornadas_row, column=2, value=f"=ROUND(C{total_row}/8,1)").number_format = '0.0 "días"'
    ws.cell(row=jornadas_row, column=2).font = font(bold=True, color=C_ROJO, size=13)
    ws.cell(row=jornadas_row, column=2).alignment = alinear("center")

    ws.cell(row=jornadas_row, column=3, value="Jornadas con Copilot").font = font(bold=True)
    ws.cell(row=jornadas_row, column=4, value=f"=ROUND(D{total_row}/8,1)").number_format = '0.0 "días"'
    ws.cell(row=jornadas_row, column=4).font = font(bold=True, color=C_VERDE, size=13)
    ws.cell(row=jornadas_row, column=4).alignment = alinear("center")

    ws.cell(row=jornadas_row, column=5, value="Jornadas ahorradas").font = font(bold=True)
    ws.cell(row=jornadas_row, column=6,
            value=f"=ROUND((C{total_row}-D{total_row})/8,1)").number_format = '0.0 "días"'
    ws.cell(row=jornadas_row, column=6).font = font(bold=True, color=C_AZUL_MED, size=13)
    ws.cell(row=jornadas_row, column=6).alignment = alinear("center")

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 28

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 4 — Modelo de datos
# ══════════════════════════════════════════════════════════════════════════════
def hoja_modelo(wb):
    ws = wb.create_sheet("Modelo de Datos")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value     = "Modelo de Datos — Entidades y Relaciones"
    ws["A1"].font      = font(bold=True, size=14, color=C_BLANCO)
    ws["A1"].fill      = fill(C_AZUL_MED)
    ws["A1"].alignment = alinear("center")
    ws.row_dimensions[1].height = 28

    entidades = {
        "TodoItem": [
            ("Id",               "int",              "PK",  "Sí", "Autogenerada"),
            ("Title",            "string",           "—",   "No", "Máx. 255 caracteres"),
            ("IsCompleted",      "bool",             "—",   "No", "Default: false"),
            ("CreatedAt",        "DateTime",         "—",   "No", "Asignada al crear"),
            ("EsRepetitiva",     "bool",             "—",   "No", "Default: false"),
            ("Recurrencia",      "TipoRecurrencia?", "—",   "Sí", "Null si no es repetitiva"),
            ("ProximaFecha",     "DateTime?",        "—",   "Sí", "Calculada al completar"),
            ("PlantillaId",      "int?",             "FK",  "Sí", "Ref. a PlantillaTarea"),
            ("UsuarioAsignadoId","int?",             "FK",  "Sí", "Ref. a UsuarioAsignado. SetNull al borrar"),
        ],
        "UsuarioAsignado": [
            ("Id",     "int",    "PK", "No", "Autogenerada"),
            ("Nombre", "string", "—",  "No", "Máx. 100 caracteres"),
            ("Email",  "string", "IX", "No", "Único, máx. 200 caracteres"),
        ],
        "PlantillaTarea": [
            ("Id",           "int",              "PK", "No",  "Autogenerada"),
            ("Titulo",       "string",           "—",  "No",  "Título por defecto"),
            ("EsRepetitiva", "bool",             "—",  "No",  "Default: false"),
            ("Recurrencia",  "TipoRecurrencia?", "—",  "Sí",  "Null si no es repetitiva"),
        ],
    }

    row = 3
    for entidad, campos in entidades.items():
        row = titulo_seccion(ws, row, f"  {entidad}", col_start=1, col_end=6)
        row = cabecera_tabla(ws, row, ["Campo", "Tipo", "Constraint", "Nullable", "Notas"], col_start=1)
        alt = False
        for campo in campos:
            fila_datos(ws, row, list(campo), col_start=1, alternada=alt)
            alt = not alt
            row += 1
        row += 1

    # Relaciones
    row = titulo_seccion(ws, row, "  Relaciones entre entidades", col_start=1, col_end=6)
    row = cabecera_tabla(ws, row, ["Origen", "Cardinalidad", "Destino", "FK", "OnDelete"], col_start=1)
    relaciones = [
        ("UsuarioAsignado", "1 → N", "TodoItem",     "UsuarioAsignadoId", "SetNull"),
        ("PlantillaTarea",  "1 → N", "TodoItem",     "PlantillaId",       "SetNull"),
        ("TodoItem",        "N → 1", "UsuarioAsignado","UsuarioAsignadoId","—"),
        ("TodoItem",        "N → 1", "PlantillaTarea", "PlantillaId",      "—"),
    ]
    alt = False
    for rel in relaciones:
        fila_datos(ws, row, list(rel), alternada=alt)
        alt = not alt
        row += 1

    anchos = {"A": 22, "B": 20, "C": 14, "D": 12, "E": 36, "F": 6}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 5 — Endpoints
# ══════════════════════════════════════════════════════════════════════════════
def hoja_endpoints(wb):
    ws = wb.create_sheet("Endpoints API")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value     = "Endpoints API REST"
    ws["A1"].font      = font(bold=True, size=14, color=C_BLANCO)
    ws["A1"].fill      = fill(C_AZUL_MED)
    ws["A1"].alignment = alinear("center")
    ws.row_dimensions[1].height = 28

    # Colores por verbo HTTP
    verbo_colores = {
        "GET":    ("61AFFE", "000000"),
        "POST":   ("49CC90", "000000"),
        "PUT":    ("FCA130", "000000"),
        "DELETE": ("F93E3E", C_BLANCO),
    }

    row = 3
    controladores = [
        ("/api/tareas", [
            ("GET",    "/api/tareas",                 "Listar todas las tareas",                                "200 + array",  "—",            "Sin parámetros"),
            ("GET",    "/api/tareas/{id}",            "Obtener tarea por ID",                                  "200 + objeto", "404",          "id: int en ruta"),
            ("POST",   "/api/tareas",                 "Crear nueva tarea",                                     "201 + objeto", "400",          "Body: CrearTareaDto"),
            ("PUT",    "/api/tareas/{id}",            "Actualizar título o estado",                            "200 + objeto", "404 / 400",    "Body: ActualizarTareaDto"),
            ("DELETE", "/api/tareas/{id}",            "Eliminar tarea permanentemente",                        "204",          "404",          "id: int en ruta"),
            ("POST",   "/api/tareas/{id}/completar",  "Completar; si repetitiva genera siguiente ocurrencia", "200 + objeto", "404",          "Sin body"),
        ]),
        ("/api/usuariosasignados", [
            ("GET",    "/api/usuariosasignados",       "Listar todos los usuarios",                           "200 + array",  "—",         "Sin parámetros"),
            ("GET",    "/api/usuariosasignados/{id}",  "Obtener usuario por ID",                              "200 + objeto", "404",       "id: int en ruta"),
            ("POST",   "/api/usuariosasignados",       "Crear usuario nuevo",                                 "201 + objeto", "400",       "Body: CrearUsuarioDto"),
            ("PUT",    "/api/usuariosasignados/{id}",  "Actualizar nombre o email",                           "200 + objeto", "404 / 400", "Body: ActualizarUsuarioDto"),
            ("DELETE", "/api/usuariosasignados/{id}",  "Eliminar (tareas quedan sin usuario)",                "204",          "404",       "Cascada SetNull en FK"),
        ]),
        ("/api/plantillas", [
            ("GET",    "/api/plantillas",                  "Listar todas las plantillas",                "200 + array",  "—",         "Sin parámetros"),
            ("GET",    "/api/plantillas/{id}",             "Obtener plantilla por ID",                   "200 + objeto", "404",       "id: int en ruta"),
            ("POST",   "/api/plantillas",                  "Crear nueva plantilla",                      "201 + objeto", "400",       "Body: CrearPlantillaDto"),
            ("PUT",    "/api/plantillas/{id}",             "Actualizar plantilla",                       "200 + objeto", "404 / 400", "Body: ActualizarPlantillaDto"),
            ("DELETE", "/api/plantillas/{id}",             "Eliminar plantilla",                         "204",          "404",       "Las tareas conservan sus valores"),
            ("POST",   "/api/plantillas/{id}/instanciar",  "Crear tarea nueva a partir de la plantilla", "201 + objeto", "404",       "Sin body; usa valores de la plantilla"),
        ]),
    ]

    for base_path, endpoints in controladores:
        row = titulo_seccion(ws, row, f"  {base_path}", col_start=1, col_end=7)
        row = cabecera_tabla(ws, row,
            ["Verbo", "Ruta completa", "Descripción", "Respuesta OK", "Error", "Notas"])
        alt = False
        for verbo, ruta, desc, ok, err, notas in endpoints:
            bg = C_AZUL_PAL if alt else C_BLANCO
            alt = not alt
            data = [verbo, ruta, desc, ok, err, notas]
            for i, v in enumerate(data):
                c = ws.cell(row=row, column=i+1, value=v)
                c.fill      = fill(bg)
                c.border    = borde_fino()
                c.alignment = alinear("center" if i in (0, 3, 4) else "left", wrap=(i == 2))
                if i == 0:
                    v_bg, v_txt = verbo_colores.get(verbo, ("CCCCCC", "000000"))
                    c.fill = fill(v_bg)
                    c.font = font(bold=True, color=v_txt, size=10)
                else:
                    c.font = font(size=10)
            ws.row_dimensions[row].height = 28
            row += 1
        row += 1

    anchos = {"A": 10, "B": 40, "C": 44, "D": 16, "E": 14, "F": 36, "G": 6}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"


# ── main ──────────────────────────────────────────────────────────────────────
def build():
    wb = Workbook()
    hoja_resumen(wb)
    hoja_estimaciones(wb)
    hoja_por_area(wb)
    hoja_modelo(wb)
    hoja_endpoints(wb)
    wb.save(OUTPUT)
    print(f"Hoja de cálculo generada: {OUTPUT}")


if __name__ == "__main__":
    build()
